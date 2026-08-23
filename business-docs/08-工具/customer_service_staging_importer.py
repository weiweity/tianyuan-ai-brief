#!/usr/bin/env python3
"""客服话术/VOC 的离线 staging 导入器。

边界：
* 只读取 Excel，输出高置信标识符已遮罩的 JSONL 与批次报告；不连接 PostgreSQL。
  这不是完整匿名化或 DLP，输出仍须位于受控私有目录或仓内 ignored output/。
* 所有输出都标记为 ``prefill``，不能直接视为正式数据或 G0 证据。
* 竞品工作表默认排除；VOC 汇总表只登记来源，行级数据来自三张明细表。

典型用法：
  python3 customer_service_staging_importer.py \
    --qa-file '/path/达肤妍产品QA&话术.xlsx' \
    --campaign-file /path/8月活动话术.xlsx \
    --voc-file /path/达肤妍核心产品VOC反馈.xlsx \
    --output-dir output/customer-service-staging-20260812

在正式数据链路准备好前，建议保留默认的每表 1000 行预览上限；完整导入
需要显式传 ``--max-rows-per-sheet 0``，仍然只写 staging 文件。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sys
import tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment diagnostic
    raise SystemExit(
        "缺少 openpyxl。请使用项目配置的 Python 运行时，或先安装 openpyxl；"
        "本工具不会自动安装依赖。"
    ) from exc

from customer_project_output_boundary import (
    ensure_sources_outside_output,
    publish_managed_directory,
    validate_output_boundary,
    write_managed_marker,
)
from customer_service_staging_contract import SOURCE_SIZE_FIELD, scrub_text


PREFILL = "prefill"
DEFAULT_BATCH_ID = "BATCH-PREFILL-20260812-001"
DEFAULT_LIMIT = 1000

VOC_DETAIL = {
    "面膜类料体问题明细": {
        "created": 0,
        "order": 1,
        "l1": 2,
        "l2": 3,
        "l3": 4,
        "product": 5,
        "desc": 6,
        "image": 7,
        "batch": 8,
        "customer": 9,
        "l4": 10,
    },
    "棉片料体问题明细": {
        "created": 0,
        "order": 1,
        "l1": 2,
        "l2": 3,
        "l3": 4,
        "product": 5,
        "desc": 6,
        "image": 7,
        "batch": 8,
    },
    "精华水|液料体问题明细": {
        "created": 0,
        "order": 1,
        "l1": 2,
        "l2": 3,
        "l3": 4,
        "product": 5,
        "desc": 6,
        "image": 7,
        "batch": 8,
    },
}

VOC_EXCLUDED = {
    "竞品分析数据源（已更新至2.13）",
    "竞品数据总览【看这个】",
    "竞品voc收集明细（已更新至3.2）",
    "竞品分析数据源（已更新至1.17）副本",
}
VOC_SUMMARY = {
    "面膜类 【周累计】",
    "棉片 【周累计】",
    "精华水|液 【周累计】-停更",
    "ALL品 全维度反馈【月】",
}

def is_present(value: Any) -> bool:
    return value not in (None, "")


def source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def iso_date(value: Any) -> str:
    return scrub_text(value, 100)


def product_family(name: Any) -> str:
    value = scrub_text(name, 240)
    if any(token in value for token in ("面膜", "冻干膜", "敷料")):
        return "面膜"
    if any(token in value for token in ("棉片", "水润贴", "水油次抛")):
        return "棉片"
    if any(token in value for token in ("精华", "精萃液", "精华水", "精研水", "喷雾", "水乳")):
        return "精华水/液"
    return "其他"


def issue_level(l1: Any, l2: Any, l3: Any, l4: Any, description: Any) -> str:
    for value in (l4, l3, l2, l1, description):
        if is_present(value):
            return scrub_text(value, 160)
    return "待人工判定"


def rows_with_numbers(worksheet: Any) -> Iterable[tuple[int, list[Any]]]:
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = list(row)
        while values and not is_present(values[-1]):
            values.pop()
        yield row_number, values


def nonempty_count(rows: list[tuple[int, list[Any]]]) -> int:
    return sum(1 for _, values in rows if any(is_present(value) for value in values))


def find_header(rows: list[tuple[int, list[Any]]], keys: tuple[str, ...]) -> tuple[int, list[Any]]:
    for row_number, values in rows:
        joined = " | ".join(scrub_text(value, 100) for value in values)
        if sum(key in joined for key in keys) >= 2:
            return row_number, values
    return (rows[0][0], rows[0][1]) if rows else (1, [])


def index_for(headers: list[Any], patterns: tuple[str, ...]) -> int | None:
    for index, header in enumerate(headers):
        text = scrub_text(header, 100)
        if any(pattern in text for pattern in patterns):
            return index
    return None


def value_at(values: list[Any], index: int | None) -> Any:
    return values[index] if index is not None and index < len(values) else ""


def make_quality() -> dict[str, Any]:
    return {
        "qa": {"source_rows": 0, "normalized_rows": 0, "preview_rows": 0, "skipped_rows": 0},
        "campaign": {"source_rows": 0, "normalized_rows": 0, "preview_rows": 0, "skipped_rows": 0},
        "voc": {
            "source_rows": 0,
            "normalized_rows": 0,
            "preview_rows": 0,
            "skipped_rows": 0,
            "duplicate_order_rows": 0,
            "image_refs_redacted": 0,
            "missing_l1": 0,
            "missing_l4": 0,
        },
        "excluded_competitor_sheets": 0,
        "voc_summary_sheets": 0,
    }


def read_workbook(path: Path):
    if not path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"只接受 .xlsx：{path}")
    # 这些飞书导出表的 XML dimension 不可靠，不能使用 read_only=True；
    # normal mode 才能完整读取实际行。预览上限控制内存与输出规模。
    return load_workbook(path, read_only=False, data_only=True)


def extract_qa(workbook: Any, file_name: str, batch_id: str, registry: list[dict[str, Any]], quality: dict[str, Any], limit: int) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    sequence = 1
    for sheet_name in workbook.sheetnames:
        rows = list(rows_with_numbers(workbook[sheet_name]))
        nonempty = nonempty_count(rows)
        if sheet_name == "目录":
            registry.append({"source_file_name": file_name, "source_sheet_name": sheet_name, "source_type": "product_qa_catalog", "included": "否", "reason": "目录/元数据，不作为记录导入", "nonempty_rows": nonempty, "normalized_rows": 0, "preview_limit": ""})
            continue
        header_number, headers = find_header(rows, ("产品", "问题", "回答", "话术"))
        product_index = index_for(headers, ("产品",))
        question_index = index_for(headers, ("业务填写问题", "问题"))
        internal_index = index_for(headers, ("品牌解答", "内部解答", "回答"))
        approved_index = index_for(headers, ("客满话术", "已审核", "快捷短语", "快捷内容", "话术"))
        product = ""
        count = 0
        skipped = 0
        for row_number, values in rows:
            if row_number <= header_number or not any(is_present(value) for value in values):
                continue
            product_value = scrub_text(value_at(values, product_index), 240)
            if product_value:
                product = product_value
            question = scrub_text(value_at(values, question_index))
            internal = scrub_text(value_at(values, internal_index))
            approved = scrub_text(value_at(values, approved_index))
            if not question and not internal and not approved:
                skipped += 1
                continue
            count += 1
            if limit and len([record for record in records if record["source_sheet_name"] == sheet_name]) >= limit:
                continue
            records.append({
                "import_batch_id": batch_id,
                "data_status": PREFILL,
                "record_id": f"QA-PREVIEW-{sequence:06d}",
                "record_type": "qa_answer",
                "record_date": "",
                "product_family": product_family(product or sheet_name),
                "product_name": product or sheet_name,
                "question_text": question,
                "internal_answer": internal,
                "approved_script": approved,
                "processing_status": "待复核",
                "review_evidence_id": "",
                "source_file_name": file_name,
                "source_sheet_name": sheet_name,
                "source_row_no": row_number,
            })
            sequence += 1
        registry.append({"source_file_name": file_name, "source_sheet_name": sheet_name, "source_type": "product_qa", "included": "是", "reason": "产品QA记录；默认每表限额", "nonempty_rows": nonempty, "normalized_rows": count, "preview_limit": limit or "完整"})
        quality["qa"]["source_rows"] += max(0, nonempty - 1)
        quality["qa"]["normalized_rows"] += count
        quality["qa"]["preview_rows"] += min(count, limit) if limit else count
        quality["qa"]["skipped_rows"] += skipped
    return records


def extract_campaign(workbook: Any, file_name: str, batch_id: str, registry: list[dict[str, Any]], quality: dict[str, Any], limit: int) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    sequence = 1
    for sheet_name in workbook.sheetnames:
        rows = list(rows_with_numbers(workbook[sheet_name]))
        nonempty = nonempty_count(rows)
        header_number, headers = find_header(rows, ("分组", "快捷", "加入"))
        date_index = index_for(headers, ("登记日期",))
        group_index = index_for(headers, ("分组", "话术分类"))
        code_index = index_for(headers, ("快捷编码",))
        content_index = index_for(headers, ("快捷短语", "快捷内容"))
        enabled_index = index_for(headers, ("是否加入团队",))
        group = ""
        record_date = ""
        count = 0
        skipped = 0
        sheet_preview_count = 0
        for row_number, values in rows:
            if row_number <= header_number or not any(is_present(value) for value in values):
                continue
            if value_at(values, group_index):
                group = scrub_text(value_at(values, group_index), 240)
            if value_at(values, date_index):
                record_date = iso_date(value_at(values, date_index))
            content = scrub_text(value_at(values, content_index))
            code = scrub_text(value_at(values, code_index), 240)
            enabled = scrub_text(value_at(values, enabled_index), 40)
            if not content and not code:
                skipped += 1
                continue
            count += 1
            if limit and sheet_preview_count >= limit:
                continue
            records.append({
                "import_batch_id": batch_id,
                "data_status": PREFILL,
                "record_id": f"CAM-PREVIEW-{sequence:06d}",
                "record_type": "campaign_script",
                "record_date": record_date,
                "group_name": group,
                "shortcut_code": code,
                "approved_script": content,
                "team_enabled": enabled,
                "processing_status": "待复核",
                "review_evidence_id": "",
                "source_file_name": file_name,
                "source_sheet_name": sheet_name,
                "source_row_no": row_number,
            })
            sequence += 1
            sheet_preview_count += 1
        registry.append({"source_file_name": file_name, "source_sheet_name": sheet_name, "source_type": "campaign_script", "included": "是", "reason": "活动话术记录；默认每表限额", "nonempty_rows": nonempty, "normalized_rows": count, "preview_limit": limit or "完整"})
        quality["campaign"]["source_rows"] += max(0, nonempty - 1)
        quality["campaign"]["normalized_rows"] += count
        quality["campaign"]["preview_rows"] += min(count, limit) if limit else count
        quality["campaign"]["skipped_rows"] += skipped
    return records


def extract_voc(workbook: Any, file_name: str, batch_id: str, registry: list[dict[str, Any]], quality: dict[str, Any], limit: int) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    sequence = 1
    for sheet_name in workbook.sheetnames:
        rows = list(rows_with_numbers(workbook[sheet_name]))
        nonempty = nonempty_count(rows)
        if sheet_name in VOC_EXCLUDED:
            registry.append({"source_file_name": file_name, "source_sheet_name": sheet_name, "source_type": "competitor", "included": "否", "reason": "按用户决定暂不纳入竞品数据", "nonempty_rows": nonempty, "normalized_rows": 0, "preview_limit": ""})
            quality["excluded_competitor_sheets"] += 1
            continue
        if sheet_name in VOC_SUMMARY:
            registry.append({"source_file_name": file_name, "source_sheet_name": sheet_name, "source_type": "voc_summary_reference", "included": "否", "reason": "汇总/透视来源，仅登记结构；行级记录以明细表为准", "nonempty_rows": nonempty, "normalized_rows": 0, "preview_limit": ""})
            quality["voc_summary_sheets"] += 1
            continue
        mapping = VOC_DETAIL.get(sheet_name)
        if not mapping:
            registry.append({"source_file_name": file_name, "source_sheet_name": sheet_name, "source_type": "other", "included": "否", "reason": "未配置映射", "nonempty_rows": nonempty, "normalized_rows": 0, "preview_limit": ""})
            continue
        count = 0
        skipped = 0
        raw_orders: list[str] = []
        image_count = 0
        sheet_preview_count = 0
        for row_number, values in rows:
            if row_number == 1 or not any(is_present(value) for value in values):
                continue
            get = lambda key: value_at(values, mapping.get(key))
            order_id = scrub_text(get("order"), 120)
            l1 = scrub_text(get("l1"), 160)
            l2 = scrub_text(get("l2"), 160)
            l3 = scrub_text(get("l3"), 160)
            l4 = scrub_text(get("l4"), 160)
            product = scrub_text(get("product"), 240)
            description = scrub_text(get("desc"), 1200)
            customer_note = scrub_text(get("customer"), 1200)
            image = scrub_text(get("image"), 200)
            batch_no = scrub_text(get("batch"), 180)
            if not any((order_id, product, l1, l2, l3, description, customer_note)):
                skipped += 1
                continue
            count += 1
            raw_orders.append(order_id)
            if image and image not in ("预览", "查看", "无", "/"):
                image_count += 1
            if limit and sheet_preview_count >= limit:
                continue
            record_id = f"VOC-PREVIEW-{sequence:06d}"
            records.append({
                "import_batch_id": batch_id,
                "data_status": PREFILL,
                "record_id": record_id,
                "record_type": "voc_feedback",
                "record_date": iso_date(get("created")),
                "product_family": product_family(product or sheet_name),
                "product_name": product,
                "order_id": f"ORDER-REDACTED-{sequence:06d}",
                "category_l1": l1,
                "category_l2": l2,
                "category_l3": l3,
                "category_l4": l4,
                "primary_issue": issue_level(l1, l2, l3, l4, description or customer_note),
                "issue_tags": "|".join(value for value in (l1, l2, l3, l4) if value),
                "question_text": description,
                "batch_no": batch_no,
                "description": customer_note or description,
                "image_ref": f"IMG-REDACTED-{sequence:06d}" if image else "",
                "feedback_count": 1,
                "escalation_level": "中",
                "processing_status": "待复核" if l1 or l2 or l3 else "待分类",
                "owner_team": "质量" if l1 == "商品问题" else "",
                "collaborating_teams": "",
                "reviewer_role": "",
                "review_evidence_id": "",
                "source_file_name": file_name,
                "source_sheet_name": sheet_name,
                "source_row_no": row_number,
            })
            sequence += 1
            sheet_preview_count += 1
        duplicate_rows = sum(count - 1 for count in Counter(order for order in raw_orders if order).values() if count > 1)
        registry.append({"source_file_name": file_name, "source_sheet_name": sheet_name, "source_type": "voc_detail", "included": "是", "reason": "VOC明细；订单号/图片引用脱敏", "nonempty_rows": nonempty, "normalized_rows": count, "preview_limit": limit or "完整"})
        quality["voc"]["source_rows"] += max(0, nonempty - 1)
        quality["voc"]["normalized_rows"] += count
        quality["voc"]["preview_rows"] += min(count, limit) if limit else count
        quality["voc"]["skipped_rows"] += skipped
        quality["voc"]["duplicate_order_rows"] += duplicate_rows
        quality["voc"]["image_refs_redacted"] += image_count
        quality["voc"]["missing_l1"] += sum(1 for record in records if record["source_sheet_name"] == sheet_name and not record["category_l1"])
        quality["voc"]["missing_l4"] += sum(1 for record in records if record["source_sheet_name"] == sheet_name and not record["category_l4"])
    return records


def build_batch(args: argparse.Namespace) -> dict[str, Any]:
    output_dir = validate_output_boundary(Path(args.output_dir))
    if output_dir.exists() and any(output_dir.iterdir()) and not args.overwrite:
        raise FileExistsError(f"输出目录已有文件，若确认覆盖请加 --overwrite：{output_dir}")
    source_paths = {
        "product_qa": Path(args.qa_file).expanduser().resolve(),
        "campaign": Path(args.campaign_file).expanduser().resolve(),
        "voc": Path(args.voc_file).expanduser().resolve(),
    }
    ensure_sources_outside_output(output_dir, source_paths.values())
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    staging_dir = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}-", dir=output_dir.parent))
    try:
        registry: list[dict[str, Any]] = []
        quality = make_quality()
        records = {
            "qa": extract_qa(read_workbook(source_paths["product_qa"]), source_paths["product_qa"].name, args.batch_id, registry, quality, args.max_rows_per_sheet),
            "campaign": extract_campaign(read_workbook(source_paths["campaign"]), source_paths["campaign"].name, args.batch_id, registry, quality, args.max_rows_per_sheet),
            "voc": extract_voc(read_workbook(source_paths["voc"]), source_paths["voc"].name, args.batch_id, registry, quality, args.max_rows_per_sheet),
        }
        for record_type, rows in records.items():
            path = staging_dir / f"{record_type}.jsonl"
            with path.open("w", encoding="utf-8") as handle:
                for row in rows:
                    handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
        files = []
        for kind, path in source_paths.items():
            workbook = read_workbook(path)
            files.append({"source_file_name": path.name, "source_type": kind, SOURCE_SIZE_FIELD: path.stat().st_size, "source_sha256": source_sha256(path), "sheet_count": len(workbook.sheetnames), "data_status": PREFILL, "raw_file_unchanged": True})
        manifest = {
            "batch_id": args.batch_id,
            "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "data_status": PREFILL,
            "preview_limit_per_sheet": args.max_rows_per_sheet or "完整",
            "source_files": files,
            "source_sheet_registry": registry,
            "quality": quality,
            "output_records": {key: len(value) for key, value in records.items()},
            "postgresql_written": False,
            "notes": [
                "这是离线 staging 输出，不是正式导入。",
                "原始 Excel 未改写；订单号、图片引用和高置信 URL/token/邮箱/手机号/身份证号已遮罩。",
                "该遮罩不是完整匿名化或 DLP；输出仍须留在受控私有目录或仓内 ignored output/。",
                "竞品工作表按用户决定暂不纳入。",
                "预填数据不计入正式 G0-03 指标。",
            ],
        }
        (staging_dir / "batch_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        (staging_dir / "quality_report.json").write_text(json.dumps(quality, ensure_ascii=False, indent=2), encoding="utf-8")
        readme = "\n".join([
            "# Customer service staging batch",
            "",
            f"- batch_id: `{args.batch_id}`",
            f"- data_status: `{PREFILL}`",
            "- PostgreSQL written: `false`",
            "- This output is prefill/staging only; it is not a publishable customer-agent release.",
            "- Review classification, evidence, duplicate-order policy, and owner routing before official import.",
            "",
        ])
        (staging_dir / "README.md").write_text(readme, encoding="utf-8")
        write_managed_marker(
            staging_dir,
            kind="customer-service-staging",
            metadata={"batch_id": args.batch_id},
        )
        publish_managed_directory(
            staging_dir,
            output_dir,
            kind="customer-service-staging",
            overwrite=args.overwrite,
        )
        return manifest
    except Exception:
        shutil.rmtree(staging_dir, ignore_errors=True)
        raise


def run_self_test() -> None:
    assert scrub_text("https://example.test/a tenant_access_token=abc") == "[URL_REDACTED] [TOKEN_REDACTED]"
    assert scrub_text("联系 13800138000 或 qa@example.test") == "联系 [PHONE_REDACTED] 或 [EMAIL_REDACTED]"
    assert scrub_text("身份证 11010519491231002X") == "身份证 [ID_REDACTED]"
    assert product_family("海葡萄面膜3.0") == "面膜"
    assert product_family("儿童棉片") == "棉片"
    assert product_family("海葡萄精萃液") == "精华水/液"
    assert issue_level("商品问题", "料体", "", "", "异物") == "料体"
    assert issue_level("", "", "", "", "") == "待人工判定"
    print("SELF_TEST_PASS")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="离线生成客服话术/VOC staging JSONL")
    parser.add_argument("--qa-file", required=False)
    parser.add_argument("--campaign-file", required=False)
    parser.add_argument("--voc-file", required=False)
    parser.add_argument("--output-dir", required=False)
    parser.add_argument("--batch-id", default=DEFAULT_BATCH_ID)
    parser.add_argument("--max-rows-per-sheet", type=int, default=DEFAULT_LIMIT, help="每张明细表最多输出多少行；0=完整")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.self_test:
        run_self_test()
        return 0
    missing = [name for name in ("qa_file", "campaign_file", "voc_file", "output_dir") if not getattr(args, name)]
    if missing:
        raise SystemExit(f"缺少参数：{', '.join('--' + name.replace('_', '-') for name in missing)}；或使用 --self-test")
    manifest = build_batch(args)
    print(json.dumps({"output_dir": args.output_dir, "batch_id": manifest["batch_id"], "output_records": manifest["output_records"], "postgresql_written": False}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, ValueError, FileExistsError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(2) from exc
