#!/usr/bin/env python3
"""只读检查售前/售后上游文件，输出不含正文的技术预填报告。"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import tempfile
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from customer_project_output_boundary import validate_output_boundary

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - 仅在实际读取 xlsx 时报告
    load_workbook = None


SCHEMA_VERSION = 1
DOMAINS = {"presale", "aftersale"}
ALLOWED_SUFFIXES = {".xlsx", ".csv"}
MAX_CELL_SCAN = 2_000_000
HEADER_SCAN_ROWS = 30

FIELD_PATTERNS = {
    "question_text": ("问题", "问法", "咨询", "客户提问", "关键词"),
    "answer_text": ("答案", "回复", "话术", "解答", "快捷短语", "快捷内容"),
    "product_context": ("产品", "商品", "sku", "品名"),
    "category_or_intent": ("分类", "类目", "场景", "意图", "问题类型"),
    "effective_from": ("开始", "生效", "上线时间"),
    "effective_to": ("结束", "失效", "截止", "下线时间"),
    "status": ("状态", "启用", "是否有效"),
    "risk": ("风险", "敏感", "禁用", "升级"),
    "updated_at": ("更新时间", "更新日期", "修改时间"),
}

SENSITIVE_PATTERNS = {
    "url": re.compile(r"(?i)(?:https?://|www\.)[^\s<>{}\[\]()\"']+"),
    "email": re.compile(r"(?i)\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b"),
    "phone": re.compile(r"(?<!\d)(?:\+?86[- ]?)?1[3-9]\d{9}(?!\d)"),
    "lark_token": re.compile(r"(?i)(?:tenant_access_token|open_id|doc_token|wiki_token|file_token)\s*[=:]\s*[^\s,;]+"),
    "lark_entity_id": re.compile(r"\b(?:ou_|oc_|on_)[A-Za-z0-9_-]{8,}\b"),
}


def sha256_file(source: Path) -> str:
    digest = hashlib.sha256()
    with source.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalized_text(value: Any, limit: int = 4000) -> str:
    if value is None:
        return ""
    return str(value).replace("\x00", "").strip()[:limit]


def recognized_fields(values: Iterable[Any]) -> list[str]:
    joined = " | ".join(normalized_text(value, 200).lower() for value in values if value is not None)
    return sorted(
        field for field, patterns in FIELD_PATTERNS.items()
        if any(pattern.lower() in joined for pattern in patterns)
    )


def scan_values(values: Iterable[Any], sensitive: Counter[str]) -> int:
    count = 0
    for value in values:
        text = normalized_text(value)
        if not text:
            continue
        count += 1
        for name, pattern in SENSITIVE_PATTERNS.items():
            sensitive[name] += len(pattern.findall(text))
    return count


def summarize_rows(rows: Iterable[tuple[Any, ...]], alias: str) -> dict[str, Any]:
    total_rows = 0
    nonempty_rows = 0
    max_columns = 0
    scanned_cells = 0
    sensitive: Counter[str] = Counter()
    header_candidates: list[tuple[int, int, int, list[str]]] = []

    for row_number, row in enumerate(rows, start=1):
        total_rows += 1
        values = tuple(row)
        max_columns = max(max_columns, len(values))
        nonempty = [value for value in values if normalized_text(value)]
        if nonempty:
            nonempty_rows += 1
        if row_number <= HEADER_SCAN_ROWS:
            fields = recognized_fields(values)
            header_candidates.append((len(fields), len(nonempty), row_number, fields))
        if scanned_cells < MAX_CELL_SCAN:
            remaining = MAX_CELL_SCAN - scanned_cells
            scanned_cells += scan_values(values[:remaining], sensitive)

    best = max(header_candidates, default=(0, 0, 0, []), key=lambda item: (item[0], item[1], -item[2]))
    return {
        "sheet_alias": alias,
        "total_rows": total_rows,
        "nonempty_rows": nonempty_rows,
        "max_columns": max_columns,
        "header_candidate_row": best[2] or None,
        "recognized_fields": best[3],
        "sensitive_hit_counts": dict(sorted(sensitive.items())),
        "scanned_nonempty_cells": scanned_cells,
        "scan_truncated": scanned_cells >= MAX_CELL_SCAN,
    }


def csv_rows(source: Path) -> Iterable[tuple[str, ...]]:
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        yield from csv.reader(handle)


def inspect_source(source: Path, domain: str) -> dict[str, Any]:
    source = source.expanduser().resolve()
    if domain not in DOMAINS:
        raise ValueError("domain 只允许 presale 或 aftersale")
    if not source.is_file():
        raise FileNotFoundError(f"输入文件不存在：{source}")
    suffix = source.suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        raise ValueError("只接受 .xlsx 或 .csv")

    sheets: list[dict[str, Any]] = []
    if suffix == ".csv":
        sheets.append(summarize_rows(csv_rows(source), "sheet_001"))
    else:
        if load_workbook is None:
            raise RuntimeError("读取 .xlsx 需要 openpyxl；请使用仓库已配置的 Python 环境")
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            for index, worksheet in enumerate(workbook.worksheets, start=1):
                sheets.append(summarize_rows(worksheet.iter_rows(values_only=True), f"sheet_{index:03d}"))
        finally:
            workbook.close()

    totals = Counter()
    for sheet in sheets:
        totals.update(sheet["sensitive_hit_counts"])
    return {
        "schema_version": SCHEMA_VERSION,
        "status": "TECHNICAL_PREFILL",
        "domain": domain,
        "source": {
            "extension": suffix,
            "size_bytes": source.stat().st_size,
            "sha256": sha256_file(source),
        },
        "sheet_count": len(sheets),
        "sheets": sheets,
        "sensitive_hit_counts": dict(sorted(totals.items())),
        "warnings": [
            "报告不包含文件名、工作表名或单元格正文；sheet_alias 仅按工作簿顺序生成。",
            "敏感命中仅是技术扫描计数，必须人工复核；不得把该报告当作 ACL、质量或内容批准 EVD。",
        ],
        "does_not_authorize": ["G0-09", "Scope#9", "G0-signature", "Ddev", "code-development", "database-write"],
    }


def write_report(report: dict[str, Any], output_dir: Path) -> Path:
    output_dir = validate_output_boundary(output_dir)
    if output_dir.exists() and any(output_dir.iterdir()):
        raise FileExistsError(f"输出目录非空，拒绝覆盖：{output_dir}")
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    temp_dir = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}-", dir=output_dir.parent))
    try:
        report_path = temp_dir / "technical_prefill.json"
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        os.replace(temp_dir, output_dir)
        return output_dir / report_path.name
    except Exception:
        for child in temp_dir.iterdir():
            child.unlink(missing_ok=True)
        temp_dir.rmdir()
        raise


def self_test() -> None:
    rows = [
        ("产品", "客户问题", "标准话术", "状态"),
        ("示例", "怎么使用", "示例回答", "启用"),
    ]
    summary = summarize_rows(iter(rows), "sheet_001")
    assert summary["header_candidate_row"] == 1
    assert {"product_context", "question_text", "answer_text", "status"}.issubset(summary["recognized_fields"])
    assert all(count == 0 for count in summary["sensitive_hit_counts"].values())
    print(json.dumps({"status": "PASS", "schema_version": SCHEMA_VERSION}, ensure_ascii=False))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="客服 Agent 售前/售后来源只读技术接收检查")
    parser.add_argument("--domain", choices=sorted(DOMAINS))
    parser.add_argument("--input")
    parser.add_argument("--output-dir")
    parser.add_argument("--self-test", action="store_true")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.self_test:
        self_test()
        return
    if not all((args.domain, args.input, args.output_dir)):
        raise SystemExit("需要 --domain、--input、--output-dir；或使用 --self-test")
    report = inspect_source(Path(args.input), args.domain)
    output = write_report(report, Path(args.output_dir))
    print(json.dumps({"status": report["status"], "domain": report["domain"], "report": str(output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
